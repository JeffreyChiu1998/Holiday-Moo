#!/usr/bin/env python3
"""
Travel Calendar Excel Export Service

This service generates Excel files with calendar view and event details for travel trips.
Requires: openpyxl, flask, flask-cors

Install dependencies:
pip install openpyxl flask flask-cors

Run the service:
python excelExportService.py
"""

import json
import sys
from datetime import datetime, timedelta
from typing import Dict, List, Any, Tuple
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import io
import base64
import urllib.parse

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

class TravelCalendarExporter:
    def __init__(self):
        self.event_colors = {
            'meeting': 'FFE6F2FF',      # Blue
            'appointment': 'FFE6F7E6',  # Green  
            'task': 'FFFFF2E6',         # Orange
            'personal': 'FFF2E6FF',     # Purple
            'travel': 'FFFFE6E6',       # Red
            'break': 'FFF5F5F5',        # Gray
            'event': 'FFFFE6F2',        # Pink
            'deadline': 'FFFFE6E6',     # Light Red
            'default': 'FFF0F8FF'       # Alice Blue
        }
        
    def create_excel_export(self, calendar_data: Dict[str, Any], trip_data: Dict[str, Any]) -> bytes:
        """Create Excel file with calendar and event list sheets"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Get trip events
        trip_events = self._get_trip_events(calendar_data['events'], trip_data['id'])
        
        # Create Calendar sheet
        calendar_sheet = wb.create_sheet("Calendar")
        custom_day_headers = calendar_data.get('customDayHeaders', {})
        self._create_calendar_sheet(calendar_sheet, trip_data, trip_events, custom_day_headers)
        
        # Create Event List sheet
        event_list_sheet = wb.create_sheet("Event List")
        self._create_event_list_sheet(event_list_sheet, trip_events, calendar_sheet)
        
        # Create Trip Summary sheet
        summary_sheet = wb.create_sheet("Trip Summary")
        self._create_trip_summary_sheet(summary_sheet, calendar_data, trip_data, trip_events)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    def _get_trip_events(self, all_events: List[Dict], trip_id: str) -> List[Dict]:
        """Filter events for the specific trip"""
        return [event for event in all_events if event.get('tripId') == trip_id]
    
    def _create_calendar_sheet(self, sheet, trip_data: Dict, events: List[Dict], custom_day_headers: Dict = None):
        """Create the calendar view sheet"""
        start_date = datetime.fromisoformat(trip_data['startDate'].replace('Z', '+00:00')).date()
        end_date = datetime.fromisoformat(trip_data['endDate'].replace('Z', '+00:00')).date()
        
        # Set sheet title
        sheet.title = "Calendar"
        
        # Create header
        sheet['A1'] = f"📅 {trip_data['name']} - Travel Calendar"
        sheet['A1'].font = Font(size=16, bold=True, color='FF1F2937')
        sheet.merge_cells('A1:H1')
        
        sheet['A2'] = f"Period: {start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
        sheet['A2'].font = Font(size=12, color='FF6B7280')
        sheet.merge_cells('A2:H2')
        
        if trip_data.get('destination'):
            sheet['A3'] = f"Destination: {trip_data['destination']}"
            sheet['A3'].font = Font(size=12, color='FF6B7280')
            sheet.merge_cells('A3:H3')
        
        # Create calendar grid
        current_row = 5
        
        # Time slots (12 AM to 11:30 PM - full 24 hours)
        time_slots = []
        for hour in range(0, 24):
            for minute in [0, 30]:
                time_str = f"{hour:02d}:{minute:02d}"
                if hour == 0:
                    display_time = f"12:{minute:02d} AM"
                elif hour < 12:
                    display_time = f"{hour}:{minute:02d} AM"
                elif hour == 12:
                    display_time = f"12:{minute:02d} PM"
                else:
                    display_time = f"{hour - 12}:{minute:02d} PM"
                time_slots.append((time_str, display_time))
        
        # Calculate date range
        current_date = start_date
        date_columns = []
        col_index = 2  # Start from column B (A is for time)
        
        while current_date <= end_date:
            date_columns.append((current_date, col_index))
            col_index += 1
            current_date += timedelta(days=1)
        
        # Create headers
        time_header_cell = sheet['A4']
        time_header_cell.value = "Time"
        time_header_cell.font = Font(bold=True, size=10, color='FFFFFFFF')
        time_header_cell.fill = PatternFill(start_color='FF3B82F6', end_color='FF3B82F6', fill_type='solid')
        time_header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for date, col in date_columns:
            cell = sheet.cell(row=4, column=col)
            
            # Check for custom day header
            date_key = date.strftime('%a %b %d %Y')  # Format: "Sat Sep 13 2025"
            custom_header = None
            
            if custom_day_headers:
                # Try different date key formats
                for key, header_data in custom_day_headers.items():
                    try:
                        # Parse the key date and compare
                        key_date = datetime.strptime(key, '%a %b %d %Y').date()
                        if key_date == date:
                            custom_header = header_data
                            break
                    except:
                        continue
            
            # Enhanced date header with custom day header if available
            date_header = f"{date.strftime('%A')}\n{date.strftime('%B %d, %Y')}"
            
            if custom_header and isinstance(custom_header, dict):
                if custom_header.get('title'):
                    date_header += f"\n🎯 {custom_header['title']}"
                if custom_header.get('description'):
                    # Truncate long descriptions
                    desc = custom_header['description']
                    if len(desc) > 30:
                        desc = desc[:30] + "..."
                    date_header += f"\n📝 {desc}"
            
            date_header += f"\n{date.strftime('%m/%d')}"
            
            cell.value = date_header
            cell.font = Font(bold=True, size=9, color='FFFFFFFF')  # White text on blue background
            cell.fill = PatternFill(start_color='FF3B82F6', end_color='FF3B82F6', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet.column_dimensions[get_column_letter(col)].width = 20
            
            # Make the header row taller to accommodate more text
            sheet.row_dimensions[4].height = 60
        
        # Create time slots
        for i, (time_str, display_time) in enumerate(time_slots):
            row = current_row + i
            sheet.cell(row=row, column=1, value=display_time)
            sheet.cell(row=row, column=1).font = Font(size=9)
            sheet.cell(row=row, column=1).fill = PatternFill(start_color='FFF9FAFB', end_color='FFF9FAFB', fill_type='solid')
        
        # Set time column width
        sheet.column_dimensions['A'].width = 12
        
        # Add events to calendar
        event_positions = {}  # Store event positions for hyperlinks
        
        for event in events:
            # Parse event times
            event_start = self._parse_local_datetime(event['startTime'])
            event_end = self._parse_local_datetime(event['endTime'])
            event_date = event_start.date()
            
            # Find the column for this date
            event_col = None
            for date, col in date_columns:
                if date == event_date:
                    event_col = col
                    break
            
            if event_col is None:
                continue
            
            # Find the row for start time
            start_time_str = event_start.strftime('%H:%M')
            start_row = None
            
            for i, (time_str, _) in enumerate(time_slots):
                if time_str <= start_time_str:
                    start_row = current_row + i
            
            if start_row is None:
                start_row = current_row
            
            # Calculate duration in 30-minute slots
            duration_minutes = (event_end - event_start).total_seconds() / 60
            duration_slots = max(1, int(duration_minutes / 30))
            
            # Get event color
            event_type = event.get('type', 'default')
            color = self.event_colors.get(event_type, self.event_colors['default'])
            
            # Create event block with sanitized data
            event_name = self._sanitize_for_excel(event.get('name', 'Event'))
            event_text = f"{event_name}\n{event_start.strftime('%H:%M')}-{event_end.strftime('%H:%M')}"
            
            event_cell = sheet.cell(row=start_row, column=event_col)
            event_cell.value = event_text
            event_cell.font = Font(size=9, bold=True, color='FF1F2937')
            event_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            event_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            event_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Store position for hyperlinks
            event_positions[event['id']] = f"{get_column_letter(event_col)}{start_row}"
            
            # Merge cells if event spans multiple time slots
            if duration_slots > 1:
                end_row = min(start_row + duration_slots - 1, current_row + len(time_slots) - 1)
                if end_row > start_row:
                    sheet.merge_cells(f"{get_column_letter(event_col)}{start_row}:{get_column_letter(event_col)}{end_row}")
        
        # Add borders to the entire calendar
        for row in range(4, current_row + len(time_slots)):
            for col in range(1, len(date_columns) + 2):
                cell = sheet.cell(row=row, column=col)
                if not cell.border.left.style:
                    cell.border = Border(
                        left=Side(style='thin', color='FFE5E7EB'),
                        right=Side(style='thin', color='FFE5E7EB'),
                        top=Side(style='thin', color='FFE5E7EB'),
                        bottom=Side(style='thin', color='FFE5E7EB')
                    )
        
        return event_positions
    
    def _create_event_list_sheet(self, sheet, events: List[Dict], calendar_sheet):
        """Create the event list sheet with hyperlinks to calendar"""
        sheet.title = "Event List"
        
        # Enhanced headers with more information
        headers = [
            "Event Name", "Date", "Day of Week", "Start Time", "End Time", "Duration", 
            "Type", "Location", "Description", "Cost", "Contact", "Tags", "Website", "Rating", "Prepaid"
        ]
        
        # Set headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(start_color='FF3B82F6', end_color='FF3B82F6', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sort events by start time
        sorted_events = sorted(events, key=lambda x: x['startTime'])
        
        # Add event data
        for row, event in enumerate(sorted_events, 2):
            try:
                # Parse event times
                event_start = self._parse_local_datetime(event['startTime'])
                event_end = self._parse_local_datetime(event['endTime'])
                
                # Calculate duration
                duration = event_end - event_start
                duration_str = f"{duration.seconds // 3600}h {(duration.seconds % 3600) // 60}m"
                
                # Event name with hyperlink to calendar
                event_name = self._sanitize_for_excel(event.get('name', 'Untitled Event'))
                name_cell = sheet.cell(row=row, column=1, value=event_name)
                name_cell.font = Font(color='FF2563EB', underline='single')
                

                
                # Other event details with sanitization
                sheet.cell(row=row, column=2, value=event_start.strftime('%Y-%m-%d'))
                sheet.cell(row=row, column=3, value=event_start.strftime('%A'))  # Day of week
                sheet.cell(row=row, column=4, value=event_start.strftime('%H:%M'))
                sheet.cell(row=row, column=5, value=event_end.strftime('%H:%M'))
                sheet.cell(row=row, column=6, value=duration_str)
                sheet.cell(row=row, column=7, value=self._sanitize_for_excel(event.get('type', 'Event')))
                
                # Handle complex location objects with hyperlinks
                location_value = event.get('location', '')
                location_str, location_url = self._extract_location_with_link(location_value)
                location_cell = sheet.cell(row=row, column=8, value=location_str)
                
                # Add hyperlink if available
                if location_url:
                    location_cell.hyperlink = location_url
                    location_cell.font = Font(color='FF2563EB', underline='single')
                
                # Description/remark
                description = self._sanitize_for_excel(event.get('remark', ''))
                sheet.cell(row=row, column=9, value=description)
                
                # Cost formatting
                cost = event.get('cost', '')
                cost_str = ""
                if cost:
                    try:
                        cost_value = float(cost)
                        cost_str = f"${cost_value:.2f}"
                    except:
                        cost_str = self._sanitize_for_excel(str(cost))
                sheet.cell(row=row, column=10, value=cost_str)
                
                # Contact
                contact = self._sanitize_for_excel(event.get('contact', ''))
                sheet.cell(row=row, column=11, value=contact)
                
                # Tags
                tags = self._sanitize_for_excel(event.get('tags', ''))
                sheet.cell(row=row, column=12, value=tags)
                
                # Website/Link
                website = self._sanitize_for_excel(event.get('link', ''))
                if website:
                    website_cell = sheet.cell(row=row, column=13, value=website)
                    website_cell.hyperlink = website
                    website_cell.font = Font(color='FF2563EB', underline='single')
                
                # Rating (if available from location data)
                rating = ""
                if isinstance(location_value, dict) and 'rating' in location_value:
                    rating = f"⭐ {location_value['rating']}"
                sheet.cell(row=row, column=14, value=rating)
                
                # Prepaid status
                prepaid_status = "✅ Yes" if event.get('isPrepaid') else "❌ No"
                sheet.cell(row=row, column=15, value=prepaid_status)
                
            except Exception as e:
                # Add a basic row with error info
                sheet.cell(row=row, column=1, value=f"Error: {event.get('name', 'Unknown Event')}")
                sheet.cell(row=row, column=2, value="Error processing event")
                continue
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 12
        
        # Special width adjustments for better readability
        sheet.column_dimensions['A'].width = 25   # Event Name
        sheet.column_dimensions['B'].width = 12   # Date
        sheet.column_dimensions['C'].width = 12   # Day of Week
        sheet.column_dimensions['D'].width = 10   # Start Time
        sheet.column_dimensions['E'].width = 10   # End Time
        sheet.column_dimensions['F'].width = 10   # Duration
        sheet.column_dimensions['G'].width = 12   # Type
        sheet.column_dimensions['H'].width = 25   # Location
        sheet.column_dimensions['I'].width = 35   # Description
        sheet.column_dimensions['J'].width = 12   # Cost
        sheet.column_dimensions['K'].width = 15   # Contact
        sheet.column_dimensions['L'].width = 20   # Tags
        sheet.column_dimensions['M'].width = 25   # Website
        sheet.column_dimensions['N'].width = 12   # Rating
        sheet.column_dimensions['O'].width = 10   # Prepaid
    
    def generate_filename(self, calendar_title: str, trip_name: str, start_date: str, end_date: str) -> str:
        """Generate filename in the specified format"""
        # Parse dates preserving original values
        start_dt = self._parse_local_datetime(start_date)
        end_dt = self._parse_local_datetime(end_date)
        
        # Format dates as yyyymmdd
        start_formatted = start_dt.strftime('%Y%m%d')
        end_formatted = end_dt.strftime('%Y%m%d')
        
        # Sanitize names for filename
        safe_title = ''.join(c for c in calendar_title if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_trip = ''.join(c for c in trip_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        
        # Replace spaces with underscores
        safe_title = safe_title.replace(' ', '_')
        safe_trip = safe_trip.replace(' ', '_')
        
        return f"{safe_title}_{safe_trip}_{start_formatted}-{end_formatted}.xlsx"
    
    def _sanitize_for_excel(self, value):
        """Sanitize data for Excel compatibility"""
        if value is None:
            return ""
        
        if isinstance(value, (dict, list)):
            return str(value)
        
        # Convert to string and limit length
        str_value = str(value)
        
        # Excel has a 32,767 character limit per cell
        if len(str_value) > 32000:
            str_value = str_value[:32000] + "..."
        
        return str_value
    
    def _extract_location_string(self, location_value):
        """Extract a readable location string from various location formats"""
        if not location_value:
            return ""
        
        if isinstance(location_value, str):
            return location_value
        
        if isinstance(location_value, dict):
            # Try different location fields in order of preference
            location_fields = ['address', 'name', 'formatted_address']
            
            for field in location_fields:
                if field in location_value and location_value[field]:
                    return str(location_value[field])
            
            # If no standard fields, try to extract coordinates
            if 'coordinates' in location_value:
                coords = location_value['coordinates']
                if isinstance(coords, dict) and 'lat' in coords and 'lng' in coords:
                    return f"Lat: {coords['lat']}, Lng: {coords['lng']}"
            
            # Fallback to string representation
            return str(location_value)
        
        return str(location_value)
    
    def _extract_location_with_link(self, location_value):
        """Extract location string and URL for hyperlinks"""
        if not location_value:
            return "", None
        
        if isinstance(location_value, str):
            return location_value, None
        
        if isinstance(location_value, dict):
            # Extract location text
            location_text = ""
            location_url = None
            
            # Try different location fields in order of preference
            location_fields = ['address', 'name', 'formatted_address']
            
            for field in location_fields:
                if field in location_value and location_value[field]:
                    location_text = str(location_value[field])
                    break
            
            # If no standard fields, try coordinates
            if not location_text and 'coordinates' in location_value:
                coords = location_value['coordinates']
                if isinstance(coords, dict) and 'lat' in coords and 'lng' in coords:
                    location_text = f"Lat: {coords['lat']}, Lng: {coords['lng']}"
            
            # Generate Google Maps URL if we have coordinates
            if 'coordinates' in location_value:
                coords = location_value['coordinates']
                if isinstance(coords, dict) and 'lat' in coords and 'lng' in coords:
                    lat = coords['lat']
                    lng = coords['lng']
                    location_url = f"https://www.google.com/maps?q={lat},{lng}"
            
            # Alternative: if we have a place ID, use that for Google Maps
            elif 'placeId' in location_value and location_value['placeId']:
                place_id = location_value['placeId']
                location_url = f"https://www.google.com/maps/place/?q=place_id:{place_id}"
            
            # Fallback: search by name/address
            elif location_text:
                encoded_location = urllib.parse.quote(location_text)
                location_url = f"https://www.google.com/maps/search/{encoded_location}"
            
            # Fallback text if nothing found
            if not location_text:
                location_text = str(location_value)
            
            return location_text, location_url
        
        return str(location_value), None
    
    def _parse_local_datetime(self, datetime_str):
        """Parse datetime string and convert to match web calendar display times"""
        if not datetime_str:
            return datetime.now()
        
        try:
            parsed_dt = None
            
            if 'T' in datetime_str:
                # ISO format with T separator
                clean_str = datetime_str.replace('Z', '').replace('+00:00', '')
                
                # Try different formats
                for fmt in ['%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M']:
                    try:
                        parsed_dt = datetime.strptime(clean_str, fmt)
                        break
                    except:
                        continue
                        
                # Fallback: use fromisoformat but remove timezone
                if parsed_dt is None:
                    parsed_dt = datetime.fromisoformat(clean_str)
            else:
                # Try standard datetime formats
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M']:
                    try:
                        parsed_dt = datetime.strptime(datetime_str, fmt)
                        break
                    except:
                        continue
            
            if parsed_dt is not None:
                # Add 8 hours to match the web calendar display
                # This converts from stored UTC time to display time
                return parsed_dt + timedelta(hours=8)
                        
        except Exception as e:
            pass
            
        # Fallback to current time
        return datetime.now()
    
    def _create_trip_summary_sheet(self, sheet, calendar_data: Dict, trip_data: Dict, events: List[Dict]):
        """Create a comprehensive trip summary sheet"""
        sheet.title = "Trip Summary"
        
        # Title
        sheet['A1'] = f"🧳 {trip_data['name']} - Trip Summary"
        sheet['A1'].font = Font(size=18, bold=True, color='FF1F2937')
        sheet.merge_cells('A1:D1')
        
        # Trip Information
        row = 3
        sheet[f'A{row}'] = "📅 Trip Information"
        sheet[f'A{row}'].font = Font(size=14, bold=True, color='FF3B82F6')
        row += 1
        
        # Parse dates for summary (preserve original times)
        start_date = self._parse_local_datetime(trip_data['startDate'])
        end_date = self._parse_local_datetime(trip_data['endDate'])
        duration = (end_date - start_date).days + 1
        
        trip_info = [
            ("Destination:", trip_data.get('destination', 'Not specified')),
            ("Start Date:", start_date.strftime('%A, %B %d, %Y')),
            ("End Date:", end_date.strftime('%A, %B %d, %Y')),
            ("Duration:", f"{duration} days"),
            ("Description:", trip_data.get('description', 'No description provided'))
        ]
        
        for label, value in trip_info:
            sheet[f'A{row}'] = label
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'] = value
            row += 1
        
        # Event Statistics
        row += 1
        sheet[f'A{row}'] = "📊 Event Statistics"
        sheet[f'A{row}'].font = Font(size=14, bold=True, color='FF3B82F6')
        row += 1
        
        # Calculate statistics
        total_events = len(events)
        event_types = {}
        total_cost = 0
        prepaid_count = 0
        events_with_location = 0
        
        for event in events:
            # Count by type
            event_type = event.get('type', 'Other')
            event_types[event_type] = event_types.get(event_type, 0) + 1
            
            # Calculate costs
            cost = event.get('cost', '')
            if cost:
                try:
                    total_cost += float(cost)
                except:
                    pass
            
            if event.get('isPrepaid'):
                prepaid_count += 1
            
            if event.get('location'):
                events_with_location += 1
        
        stats = [
            ("Total Events:", str(total_events)),
            ("Events with Location:", f"{events_with_location} ({events_with_location/total_events*100:.1f}%)" if total_events > 0 else "0"),
            ("Total Estimated Cost:", f"${total_cost:.2f}" if total_cost > 0 else "Not specified"),
            ("Prepaid Events:", f"{prepaid_count} ({prepaid_count/total_events*100:.1f}%)" if total_events > 0 else "0"),
        ]
        
        for label, value in stats:
            sheet[f'A{row}'] = label
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'] = value
            row += 1
        
        # Event Types Breakdown
        if event_types:
            row += 1
            sheet[f'A{row}'] = "🏷️ Event Types"
            sheet[f'A{row}'].font = Font(size=14, bold=True, color='FF3B82F6')
            row += 1
            
            for event_type, count in sorted(event_types.items()):
                percentage = (count / total_events * 100) if total_events > 0 else 0
                sheet[f'A{row}'] = f"{event_type.title()}:"
                sheet[f'A{row}'].font = Font(bold=True)
                sheet[f'B{row}'] = f"{count} events ({percentage:.1f}%)"
                row += 1
        
        # Daily Schedule Overview
        row += 1
        sheet[f'A{row}'] = "📅 Daily Schedule Overview"
        sheet[f'A{row}'].font = Font(size=14, bold=True, color='FF3B82F6')
        row += 1
        
        # Group events by date
        events_by_date = {}
        for event in events:
            event_date = self._parse_local_datetime(event['startTime']).date()
            if event_date not in events_by_date:
                events_by_date[event_date] = []
            events_by_date[event_date].append(event)
        
        # Create daily overview
        current_date = start_date.date()
        while current_date <= end_date.date():
            day_events = events_by_date.get(current_date, [])
            sheet[f'A{row}'] = current_date.strftime('%A, %B %d')
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'] = f"{len(day_events)} events"
            
            if day_events:
                event_names = [event.get('name', 'Untitled') for event in day_events[:3]]
                if len(day_events) > 3:
                    event_names.append(f"... and {len(day_events) - 3} more")
                sheet[f'C{row}'] = ", ".join(event_names)
            
            row += 1
            current_date += timedelta(days=1)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 40
        sheet.column_dimensions['D'].width = 15

# Flask API endpoints
exporter = TravelCalendarExporter()

@app.route('/export-trip', methods=['POST'])
def export_trip():
    """Export trip calendar as Excel file"""
    try:
        data = request.json
        calendar_data = data.get('calendarData')
        trip_data = data.get('tripData')
        
        if not calendar_data or not trip_data:
            return jsonify({'error': 'Missing calendar or trip data'}), 400
        
        # Generate Excel file
        excel_bytes = exporter.create_excel_export(calendar_data, trip_data)
        
        # Generate filename
        filename = exporter.generate_filename(
            calendar_data.get('title', 'Calendar'),
            trip_data['name'],
            trip_data['startDate'],
            trip_data['endDate']
        )
        
        # Convert to base64 for JSON response
        excel_b64 = base64.b64encode(excel_bytes).decode('utf-8')
        
        return jsonify({
            'success': True,
            'filename': filename,
            'data': excel_b64
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'service': 'excel-export'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=False)