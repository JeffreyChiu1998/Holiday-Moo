#!/usr/bin/env python3
"""
Holiday Moo - Local Excel Export Service
Beautiful calendar-focused Excel dashboard generator
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import base64
import io
from datetime import datetime, timedelta
import json

app = Flask(__name__)
CORS(app)

class HolidayMooExcelGenerator:
    def __init__(self):
        # Color scheme for professional dashboard
        self.colors = {
            'header': 'FF2E4057',      # Dark blue-gray
            'primary': 'FF3498DB',     # Blue
            'secondary': 'FF2ECC71',   # Green
            'accent': 'FFE67E22',      # Orange
            'warning': 'FFF39C12',     # Yellow
            'danger': 'FFE74C3C',      # Red
            'light': 'FFECF0F1',       # Light gray
            'white': 'FFFFFFFF',       # White
            'text': 'FF2C3E50',        # Dark text
        }
        
        # Event type colors
        self.event_colors = {
            'dining': 'FFE67E22',      # Orange
            'sightseeing': 'FF3498DB', # Blue
            'transport': 'FF9B59B6',   # Purple
            'accommodation': 'FF2ECC71', # Green
            'activity': 'FFF39C12',    # Yellow
            'shopping': 'FFE91E63',    # Pink
            'default': 'FF95A5A6'      # Gray
        }
        
        # Time slots for calendar (30-minute intervals)
        self.time_slots = []
        for hour in range(6, 24):  # 6:00 AM to 11:30 PM
            self.time_slots.append(f"{hour:02d}:00")
            self.time_slots.append(f"{hour:02d}:30")

    def create_workbook(self, calendar_data, trip_data):
        """Create the main workbook with all sheets"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets in order
        self.create_calendar_sheet(wb, calendar_data, trip_data)
        self.create_overview_sheet(wb, calendar_data, trip_data)
        self.create_events_sheet(wb, calendar_data, trip_data)
        self.create_summary_sheet(wb, calendar_data, trip_data)
        
        return wb

    def create_calendar_sheet(self, wb, calendar_data, trip_data):
        """Create the main calendar dashboard sheet"""
        ws = wb.create_sheet("📅 Trip Calendar", 0)
        
        # Get trip dates and events
        start_date = self.parse_datetime(trip_data.get('startDate', '2025-01-01'))
        end_date = self.parse_datetime(trip_data.get('endDate', '2025-01-02'))
        trip_events = self.get_trip_events(calendar_data.get('events', []), start_date, end_date)
        
        # Calculate trip duration and create date range
        duration = (end_date - start_date).days + 1
        dates = [start_date + timedelta(days=i) for i in range(duration)]
        
        # Create header section
        self.create_calendar_header(ws, trip_data, dates)
        
        # Create calendar grid
        self.create_calendar_grid(ws, dates, trip_events)
        
        # Add legend
        self.create_calendar_legend(ws, len(dates))
        
        # Set column widths and row heights
        self.format_calendar_sheet(ws, len(dates))

    def create_calendar_header(self, ws, trip_data, dates):
        """Create the header section with trip info"""
        # Main title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"🏖️ {trip_data.get('name', 'Holiday Moo Trip')}"
        title_cell.font = Font(size=20, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Trip details
        ws.merge_cells('A2:H2')
        details_cell = ws['A2']
        start_str = dates[0].strftime('%B %d, %Y')
        end_str = dates[-1].strftime('%B %d, %Y')
        details_cell.value = f"📍 {trip_data.get('destination', 'Unknown')} • {start_str} - {end_str} • {len(dates)} days"
        details_cell.font = Font(size=12, color='FFFFFF')
        details_cell.fill = PatternFill(start_color=self.colors['primary'], end_color=self.colors['primary'], fill_type='solid')
        details_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set row heights
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 10  # Spacer

    def create_calendar_grid(self, ws, dates, events):
        """Create the main calendar grid with time slots and events"""
        start_row = 4
        
        # Create day headers
        ws.cell(row=start_row, column=1, value="Time")
        ws.cell(row=start_row, column=1).font = Font(bold=True, color='FFFFFF')
        ws.cell(row=start_row, column=1).fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        
        for i, date in enumerate(dates, 2):
            day_cell = ws.cell(row=start_row, column=i)
            day_cell.value = f"{date.strftime('%a')}\n{date.strftime('%m/%d')}"
            day_cell.font = Font(bold=True, color='FFFFFF', size=10)
            day_cell.fill = PatternFill(start_color=self.colors['primary'], end_color=self.colors['primary'], fill_type='solid')
            day_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Create time slots
        for i, time_slot in enumerate(self.time_slots, start_row + 1):
            time_cell = ws.cell(row=i, column=1, value=time_slot)
            time_cell.font = Font(size=9, color=self.colors['text'])
            time_cell.fill = PatternFill(start_color=self.colors['light'], end_color=self.colors['light'], fill_type='solid')
            time_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Create empty cells for each day
            for j in range(2, len(dates) + 2):
                cell = ws.cell(row=i, column=j, value="")
                cell.fill = PatternFill(start_color=self.colors['white'], end_color=self.colors['white'], fill_type='solid')
        
        # Place events in calendar
        self.place_events_in_calendar(ws, dates, events, start_row)
        
        # Add borders to calendar grid
        self.add_calendar_borders(ws, start_row, len(dates), len(self.time_slots))

    def extract_date_simple(self, date_string):
        """Simple string-based date extraction - no datetime conversion"""
        if not date_string:
            return "2025-01-01"
            
        date_str = str(date_string).strip()
        print(f"🗓️ Raw date input: '{date_str}'")
        
        # Extract YYYY-MM-DD pattern
        import re
        match = re.search(r'(\d{4})-(\d{2})-(\d{2})', date_str)
        if match:
            result = f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
            print(f"✅ Date extracted: '{result}'")
            return result
        
        print(f"⚠️ No date found in '{date_str}', using default")
        return "2025-01-01"

    def extract_date_only(self, datetime_string):
        """Extract date as datetime.date object"""
        date_str = self.extract_date_simple(datetime_string)
        parts = date_str.split('-')
        year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
        return datetime(year, month, day).date()

    def place_events_in_calendar(self, ws, dates, events, start_row):
        """Place events in their appropriate time slots with cell merging"""
        # Track merged cells to avoid conflicts
        merged_cells = set()
        
        for event in events:
            try:
                # Extract date and times separately to avoid timezone issues
                event_date = self.extract_date_only(event.get('startTime', '2025-01-01'))
                print(f"📅 Event date: {event_date}")
            except Exception as e:
                print(f"Warning: Could not parse event time {event.get('startTime', 'unknown')}: {e}")
                continue
            
            # Find the date column
            date_col = None
            for i, date in enumerate(dates):
                if date.date() == event_date:
                    date_col = i + 2  # +2 because column 1 is time, dates start at column 2
                    break
            
            if date_col is None:
                continue
            
            # Find start and end time slots using simple string extraction
            start_time_slot = self.find_time_slot_simple(event.get('startTime', '2025-01-01T09:00:00'))
            end_time_slot = self.find_time_slot_simple(event.get('endTime', event.get('startTime', '2025-01-01T10:00:00')))
            
            if start_time_slot is None:
                continue
            
            # Calculate rows
            start_row_idx = start_row + 1 + start_time_slot
            end_row_idx = start_row + 1 + end_time_slot if end_time_slot is not None else start_row_idx
            
            # Ensure end row is at least one slot after start
            if end_row_idx <= start_row_idx:
                end_row_idx = start_row_idx + 1
            
            # Format event text - try multiple fields for event name
            event_title = event.get('title') or event.get('name') or event.get('eventName') or 'Untitled Event'
            event_title = self.safe_excel_value(event_title)
            event_text = f"{event_title}"
            
            # Add location info
            location_value = event.get('location')
            if location_value:
                if isinstance(location_value, dict):
                    location_name = location_value.get('name', '')
                    if location_name:
                        event_text += f"\n📍 {location_name}"
                else:
                    location_str = self.safe_excel_value(location_value)
                    event_text += f"\n📍 {location_str}"
            
            # Get event color based on type
            event_type = event.get('type', 'default').lower()
            event_color = self.event_colors.get(event_type, self.event_colors['default'])
            
            # Merge cells if event spans multiple time slots
            if end_row_idx > start_row_idx:
                try:
                    # Validate merge range before attempting
                    if start_row_idx > 0 and end_row_idx > start_row_idx and date_col > 0:
                        start_cell = f"{get_column_letter(date_col)}{start_row_idx}"
                        end_cell = f"{get_column_letter(date_col)}{end_row_idx - 1}"
                        
                        # Check if cells are already merged or contain data
                        merge_range = f"{start_cell}:{end_cell}"
                        print(f"🔗 Attempting to merge: {merge_range}")
                        
                        # Check for conflicts with existing merges
                        conflict = False
                        for row in range(start_row_idx, end_row_idx):
                            cell_key = f"{date_col}_{row}"
                            if cell_key in merged_cells:
                                print(f"⚠️ Merge conflict detected at {get_column_letter(date_col)}{row}")
                                conflict = True
                                break
                        
                        if not conflict:
                            # Clear any existing content in the range first
                            for row in range(start_row_idx, end_row_idx):
                                cell_to_clear = ws.cell(row=row, column=date_col)
                                if cell_to_clear.value:
                                    print(f"⚠️ Clearing existing content in {get_column_letter(date_col)}{row}")
                                    cell_to_clear.value = None
                            
                            # Perform the merge
                            ws.merge_cells(merge_range)
                            print(f"✅ Successfully merged: {merge_range}")
                            
                            # Track merged cells
                            for row in range(start_row_idx, end_row_idx):
                                merged_cells.add(f"{date_col}_{row}")
                        else:
                            print(f"❌ Skipping merge due to conflict: {merge_range}")
                            raise ValueError("Merge conflict")
                        
                        # Set value and formatting on the merged cell
                        cell = ws.cell(row=start_row_idx, column=date_col)
                        cell.value = event_text
                        cell.font = Font(size=8, bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color=event_color, end_color=event_color, fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                    else:
                        print(f"⚠️ Invalid merge range: start_row={start_row_idx}, end_row={end_row_idx}, col={date_col}")
                        raise ValueError("Invalid merge range")
                        
                except Exception as e:
                    print(f"❌ Could not merge cells for event {event_title}: {e}")
                    # Fallback to single cell
                    cell = ws.cell(row=start_row_idx, column=date_col)
                    cell.value = event_text
                    cell.font = Font(size=8, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=event_color, end_color=event_color, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                # Single cell event
                cell = ws.cell(row=start_row_idx, column=date_col)
                cell.value = event_text
                cell.font = Font(size=8, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=event_color, end_color=event_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def extract_time_simple(self, time_string):
        """Simple string-based time extraction - no datetime conversion at all"""
        if not time_string:
            return "09:00"
            
        # Convert to string and clean it
        time_str = str(time_string).strip()
        print(f"🔍 Raw time input: '{time_str}'")
        
        # Method 1: Direct regex extraction (timezone-naive)
        import re
        
        # Look for HH:MM pattern after T - treat as local time
        match = re.search(r'T(\d{1,2}):(\d{2})', time_str)
        if match:
            hour = int(match.group(1))
            minute = int(match.group(2))
            result = f"{hour:02d}:{minute:02d}"
            print(f"✅ Local time extracted: '{result}' (no timezone conversion)")
            return result
        
        # Method 2: Manual string splitting
        if 'T' in time_str:
            try:
                parts = time_str.split('T')
                if len(parts) > 1:
                    time_part = parts[1]
                    # Remove everything after the time (Z, +00:00, etc.)
                    time_part = time_part.split('Z')[0].split('+')[0].split('-')[0].split('.')[0]
                    
                    if ':' in time_part:
                        time_components = time_part.split(':')
                        if len(time_components) >= 2:
                            hour = int(time_components[0])
                            minute = int(time_components[1])
                            result = f"{hour:02d}:{minute:02d}"
                            print(f"✅ Manual extracted: '{result}'")
                            return result
            except Exception as e:
                print(f"⚠️ Manual extraction error: {e}")
        
        # Method 3: Look for any HH:MM pattern
        match = re.search(r'(\d{1,2}):(\d{2})', time_str)
        if match:
            hour = int(match.group(1))
            minute = int(match.group(2))
            result = f"{hour:02d}:{minute:02d}"
            print(f"✅ General pattern extracted: '{result}'")
            return result
        
        print(f"⚠️ No time found in '{time_str}', using default 09:00")
        return "09:00"

    def extract_local_time(self, datetime_string):
        """Extract time components as integers"""
        time_str = self.extract_time_simple(datetime_string)
        parts = time_str.split(':')
        return int(parts[0]), int(parts[1])

    def find_time_slot_simple(self, time_string):
        """Find time slot using simple string extraction"""
        time_str = self.extract_time_simple(time_string)
        print(f"🎯 Finding slot for time: '{time_str}'")
        
        # Parse the time string
        parts = time_str.split(':')
        hour = int(parts[0])
        minute = int(parts[1])
        
        # Round to nearest 30-minute slot
        if minute < 15:
            target_time = f"{hour:02d}:00"
        elif minute < 45:
            target_time = f"{hour:02d}:30"
        else:
            # Round up to next hour
            next_hour = (hour + 1) % 24
            target_time = f"{next_hour:02d}:00"
        
        print(f"🎯 Rounded {time_str} -> {target_time}")
        
        # Find the slot index
        try:
            slot_index = self.time_slots.index(target_time)
            print(f"✅ Found slot index: {slot_index}")
            return slot_index
        except ValueError:
            # If not found, find closest slot
            for i, slot in enumerate(self.time_slots):
                if slot >= target_time:
                    print(f"✅ Closest slot index: {i}")
                    return i
            print(f"✅ Using last slot: {len(self.time_slots) - 1}")
            return len(self.time_slots) - 1  # Last slot if nothing found

    def find_time_slot(self, datetime_obj):
        """Wrapper for backward compatibility"""
        if isinstance(datetime_obj, str):
            return self.find_time_slot_simple(datetime_obj)
        else:
            # Convert datetime to string format and process
            time_str = f"{datetime_obj.hour:02d}:{datetime_obj.minute:02d}"
            return self.find_time_slot_simple(f"T{time_str}:00")



    def add_calendar_borders(self, ws, start_row, num_days, num_time_slots):
        """Add professional borders to the calendar grid"""
        thin_border = Border(
            left=Side(style='thin', color='FF000000'),
            right=Side(style='thin', color='FF000000'),
            top=Side(style='thin', color='FF000000'),
            bottom=Side(style='thin', color='FF000000')
        )
        
        thick_border = Border(
            left=Side(style='thick', color='FF000000'),
            right=Side(style='thick', color='FF000000'),
            top=Side(style='thick', color='FF000000'),
            bottom=Side(style='thick', color='FF000000')
        )
        
        # Apply borders to entire calendar grid
        for row in range(start_row, start_row + num_time_slots + 1):
            for col in range(1, num_days + 2):
                cell = ws.cell(row=row, column=col)
                if row == start_row:  # Header row
                    cell.border = thick_border
                else:
                    cell.border = thin_border

    def create_calendar_legend(self, ws, num_days):
        """Create legend for event types"""
        legend_start_row = 4 + len(self.time_slots) + 3
        
        # Legend title - moved one column right (column 2 instead of 1)
        ws.cell(row=legend_start_row, column=2, value="Event Types:")
        ws.cell(row=legend_start_row, column=2).font = Font(bold=True, size=12)
        
        # Legend items
        legend_items = [
            ('🍽️ Dining', 'dining'),
            ('🏛️ Sightseeing', 'sightseeing'),
            ('🚗 Transport', 'transport'),
            ('🏨 Accommodation', 'accommodation'),
            ('🎯 Activity', 'activity'),
            ('🛍️ Shopping', 'shopping')
        ]
        
        for i, (label, event_type) in enumerate(legend_items):
            row = legend_start_row + 1 + (i // 3)
            col = 2 + (i % 3) * 2  # Start from column 2 instead of 1
            
            # Color box
            color_cell = ws.cell(row=row, column=col)
            color_cell.value = "  "
            color_cell.fill = PatternFill(
                start_color=self.event_colors[event_type], 
                end_color=self.event_colors[event_type], 
                fill_type='solid'
            )
            
            # Label
            label_cell = ws.cell(row=row, column=col + 1, value=label)
            label_cell.font = Font(size=9)

    def format_calendar_sheet(self, ws, num_days):
        """Set column widths and row heights for calendar"""
        # Time column
        ws.column_dimensions['A'].width = 10
        
        # Date columns - much wider for better event display
        for i in range(2, num_days + 2):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 25  # Increased from 15 to 25
        
        # Row heights for time slots (smaller since we have 30-minute slots)
        for i in range(5, 5 + len(self.time_slots)):
            ws.row_dimensions[i].height = 25

    def create_overview_sheet(self, wb, calendar_data, trip_data):
        """Create comprehensive trip analytics and overview sheet"""
        ws = wb.create_sheet("📊 Trip Analytics")
        
        # Header
        ws.merge_cells('A1:H1')
        ws['A1'].value = f"📊 Trip Analytics Dashboard - {trip_data.get('name', 'Holiday Moo Trip')}"
        ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        # Trip statistics
        start_date = self.parse_datetime(trip_data.get('startDate', '2025-01-01'))
        end_date = self.parse_datetime(trip_data.get('endDate', '2025-01-02'))
        duration = (end_date - start_date).days + 1
        trip_events = self.get_trip_events(calendar_data.get('events', []), start_date, end_date)
        
        # Calculate financial analytics
        total_cost, budget, cost_breakdown = self.calculate_financial_analytics(trip_events, trip_data)
        
        # Basic trip info section
        ws.merge_cells('A3:D3')
        ws['A3'].value = "🎯 Trip Summary"
        ws['A3'].font = Font(size=14, bold=True, color=self.colors['text'])
        ws['A3'].fill = PatternFill(start_color=self.colors['secondary'], end_color=self.colors['secondary'], fill_type='solid')
        
        basic_stats = [
            ('📍 Destination', self.safe_excel_value(trip_data.get('destination', 'Unknown'))),
            ('📅 Start Date', start_date.strftime('%B %d, %Y')),
            ('📅 End Date', end_date.strftime('%B %d, %Y')),
            ('⏱️ Duration', f"{duration} days"),
            ('📋 Total Events', len(trip_events)),
            ('📊 Events per Day', f"{len(trip_events) / duration:.1f}"),
        ]
        
        for i, (label, value) in enumerate(basic_stats, 4):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=self.safe_excel_value(value))
        
        # Financial analytics section
        ws.merge_cells('A11:D11')
        ws['A11'].value = "💰 Financial Analytics"
        ws['A11'].font = Font(size=14, bold=True, color=self.colors['text'])
        ws['A11'].fill = PatternFill(start_color=self.colors['accent'], end_color=self.colors['accent'], fill_type='solid')
        
        budget_over_under = "Over Budget" if total_cost > budget else "Under Budget" if total_cost < budget else "On Budget"
        budget_percentage = (total_cost / budget * 100) if budget > 0 else 0
        
        financial_stats = [
            ('💵 Total Estimated Cost', f"${total_cost:.2f}"),
            ('🎯 Budget', f"${budget:.2f}"),
            ('📊 Budget Usage', f"{budget_percentage:.1f}%"),
            ('⚖️ Budget Status', budget_over_under),
            ('💸 Cost per Day', f"${total_cost / duration:.2f}"),
            ('🎫 Average Event Cost', f"${total_cost / len(trip_events):.2f}" if trip_events else "$0.00"),
        ]
        
        for i, (label, value) in enumerate(financial_stats, 12):
            cell = ws.cell(row=i, column=1, value=label)
            cell.font = Font(bold=True)
            value_cell = ws.cell(row=i, column=2, value=value)
            
            # Color code budget status
            if "Over Budget" in str(value):
                value_cell.fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
                value_cell.font = Font(color='FF9C0006', bold=True)
            elif "Under Budget" in str(value):
                value_cell.fill = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')
                value_cell.font = Font(color='FF006100', bold=True)
        
        # Event type analytics
        ws.merge_cells('F3:H3')
        ws['F3'].value = "📊 Event Type Distribution"
        ws['F3'].font = Font(size=14, bold=True, color=self.colors['text'])
        ws['F3'].fill = PatternFill(start_color=self.colors['primary'], end_color=self.colors['primary'], fill_type='solid')
        
        event_types = {}
        for event in trip_events:
            event_type = event.get('type', 'Other')
            event_types[event_type] = event_types.get(event_type, 0) + 1
        
        # Event type breakdown with percentages
        ws.cell(row=4, column=6, value="Event Type").font = Font(bold=True)
        ws.cell(row=4, column=7, value="Count").font = Font(bold=True)
        ws.cell(row=4, column=8, value="Percentage").font = Font(bold=True)
        
        for i, (event_type, count) in enumerate(event_types.items(), 5):
            percentage = (count / len(trip_events) * 100) if trip_events else 0
            ws.cell(row=i, column=6, value=event_type)
            ws.cell(row=i, column=7, value=count)
            ws.cell(row=i, column=8, value=f"{percentage:.1f}%")
        
        # Cost breakdown by category
        ws.merge_cells('F12:H12')
        ws['F12'].value = "💰 Cost Breakdown by Category"
        ws['F12'].font = Font(size=14, bold=True, color=self.colors['text'])
        ws['F12'].fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
        
        ws.cell(row=13, column=6, value="Category").font = Font(bold=True)
        ws.cell(row=13, column=7, value="Cost").font = Font(bold=True)
        ws.cell(row=13, column=8, value="% of Total").font = Font(bold=True)
        
        for i, (category, cost) in enumerate(cost_breakdown.items(), 14):
            cost_percentage = (cost / total_cost * 100) if total_cost > 0 else 0
            ws.cell(row=i, column=6, value=category)
            ws.cell(row=i, column=7, value=f"${cost:.2f}")
            ws.cell(row=i, column=8, value=f"{cost_percentage:.1f}%")
        
        # Daily activity analysis
        ws.merge_cells('A20:H20')
        ws['A20'].value = "📅 Daily Activity Analysis"
        ws['A20'].font = Font(size=14, bold=True, color=self.colors['text'])
        ws['A20'].fill = PatternFill(start_color=self.colors['secondary'], end_color=self.colors['secondary'], fill_type='solid')
        
        # Group events by day
        daily_events = {}
        daily_costs = {}
        for event in trip_events:
            event_date = self.parse_datetime(event.get('startTime', '2025-01-01T00:00:00')).date()
            day_key = event_date.strftime('%m/%d')
            
            if day_key not in daily_events:
                daily_events[day_key] = 0
                daily_costs[day_key] = 0
            
            daily_events[day_key] += 1
            event_cost = self.extract_cost_value(event.get('cost', event.get('estimatedCost', 0)))
            daily_costs[day_key] += event_cost
        
        # Daily breakdown headers
        headers = ['Date', 'Day of Week', 'Events', 'Total Cost', 'Avg Cost/Event']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=21, column=i, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        
        # Daily data
        current_date = start_date
        for i in range(duration):
            day_key = current_date.strftime('%m/%d')
            events_count = daily_events.get(day_key, 0)
            day_cost = daily_costs.get(day_key, 0)
            avg_cost = day_cost / events_count if events_count > 0 else 0
            
            row = 22 + i
            ws.cell(row=row, column=1, value=current_date.strftime('%m/%d/%Y'))
            ws.cell(row=row, column=2, value=current_date.strftime('%A'))
            ws.cell(row=row, column=3, value=events_count)
            ws.cell(row=row, column=4, value=f"${day_cost:.2f}")
            ws.cell(row=row, column=5, value=f"${avg_cost:.2f}")
            
            # Color code high activity days
            if events_count >= 4:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFEB9C', end_color='FFFFEB9C', fill_type='solid')
            
            current_date += timedelta(days=1)
        
        # Set column widths
        column_widths = [20, 25, 15, 15, 20, 20, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def calculate_financial_analytics(self, events, trip_data):
        """Calculate comprehensive financial analytics"""
        total_cost = 0
        cost_breakdown = {}
        
        # Extract budget from trip data
        budget = trip_data.get('budget', 0)
        if isinstance(budget, str):
            budget = self.extract_cost_value(budget)
        
        # If no budget set, estimate based on duration and destination
        if budget == 0:
            duration = 7  # Default
            try:
                start_date = self.parse_datetime(trip_data.get('startDate', '2025-01-01'))
                end_date = self.parse_datetime(trip_data.get('endDate', '2025-01-02'))
                duration = (end_date - start_date).days + 1
            except:
                pass
            budget = duration * 150  # $150 per day default budget
        
        # Calculate costs by category
        for event in events:
            event_cost = self.extract_cost_value(event.get('cost', event.get('estimatedCost', 0)))
            total_cost += event_cost
            
            event_type = event.get('type', 'Other')
            if event_type not in cost_breakdown:
                cost_breakdown[event_type] = 0
            cost_breakdown[event_type] += event_cost
        
        return total_cost, budget, cost_breakdown

    def extract_cost_value(self, cost_info):
        """Extract numeric cost value from various formats"""
        if not cost_info or cost_info in ['N/A', 'NaN', None, 'Free', 'free']:
            return 0
        
        if isinstance(cost_info, (int, float)):
            return float(cost_info)
        
        if isinstance(cost_info, str):
            # Remove currency symbols and extract numbers
            import re
            numbers = re.findall(r'\d+\.?\d*', cost_info.replace(',', ''))
            if numbers:
                try:
                    return float(numbers[0])
                except ValueError:
                    pass
        
        return 0

    def create_events_sheet(self, wb, calendar_data, trip_data):
        """Create detailed events list sheet with comprehensive information"""
        ws = wb.create_sheet("📋 Events Details")
        
        # Header
        headers = ['#', 'Date', 'Start Time', 'End Time', 'Event Name', 'Location', 'Address', 'Type', 'Cost', 'Paid Status', 'Description', 'Notes']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Get trip events
        start_date = self.parse_datetime(trip_data.get('startDate', '2025-01-01'))
        end_date = self.parse_datetime(trip_data.get('endDate', '2025-01-02'))
        trip_events = self.get_trip_events(calendar_data.get('events', []), start_date, end_date)
        
        # Sort events by date and time
        trip_events.sort(key=lambda x: self.parse_datetime(x.get('startTime', '2025-01-01T00:00:00')))
        
        # Add events
        for i, event in enumerate(trip_events, 2):
            # Extract date and times separately to avoid timezone conversion
            event_date = self.extract_date_only(event.get('startTime', '2025-01-01'))
            start_hour, start_minute = self.extract_local_time(event.get('startTime', '2025-01-01T09:00:00'))
            end_hour, end_minute = self.extract_local_time(event.get('endTime', event.get('startTime', '2025-01-01T10:00:00')))
            
            # Extract location information
            location_info = event.get('location', {})
            location_name = ""
            location_address = ""
            location_url = ""
            
            if isinstance(location_info, dict):
                location_name = location_info.get('name', 'TBD')
                location_address = location_info.get('address', '')
                # Create Google Maps URL if coordinates available
                coords = location_info.get('coordinates', {})
                if coords and 'lat' in coords and 'lng' in coords:
                    location_url = f"https://www.google.com/maps?q={coords['lat']},{coords['lng']}"
            else:
                location_name = str(location_info) if location_info else 'TBD'
            
            # Extract cost information
            cost_info = event.get('cost', event.get('estimatedCost', 'N/A'))
            cost_display = self.format_cost(cost_info)
            
            # Determine paid status
            paid_status = self.get_paid_status(event)
            
            # Get proper event title
            event_title = event.get('title', event.get('name', 'Untitled Event'))
            
            # Use simple string extraction for display
            start_time_str = self.extract_time_simple(event.get('startTime', '09:00'))
            end_time_str = self.extract_time_simple(event.get('endTime', event.get('startTime', '10:00')))
            date_str = self.extract_date_simple(event.get('startTime', '2025-01-01'))
            
            # Convert date string to display format
            date_parts = date_str.split('-')
            display_date = f"{date_parts[1]}/{date_parts[2]}/{date_parts[0]}"
            
            # Row data with pure string-based times (no conversion)
            ws.cell(row=i, column=1, value=i-1)  # #
            ws.cell(row=i, column=2, value=display_date)  # Date (string-based)
            ws.cell(row=i, column=3, value=start_time_str)  # Start Time (string-based)
            ws.cell(row=i, column=4, value=end_time_str)  # End Time (string-based)
            ws.cell(row=i, column=5, value=self.safe_excel_value(event_title))  # Event Name
            
            # Location with hyperlink if available
            location_cell = ws.cell(row=i, column=6, value=self.safe_excel_value(location_name))
            if location_url:
                location_cell.hyperlink = location_url
                location_cell.font = Font(color='FF0000FF', underline='single')  # Blue underlined
            
            ws.cell(row=i, column=7, value=self.safe_excel_value(location_address))  # Address
            ws.cell(row=i, column=8, value=self.safe_excel_value(event.get('type', 'Event')))  # Type
            ws.cell(row=i, column=9, value=cost_display)  # Cost
            ws.cell(row=i, column=10, value=paid_status)  # Paid Status
            # Get description from multiple possible fields with enhanced debug logging
            description_fields = ['remark', 'description', 'desc', 'details', 'note', 'notes', 'summary', 'content', 'body', 'text']
            description = ''
            
            print(f"\n🔍 Processing event: '{event.get('title', event.get('name', 'Unknown'))}'")
            print(f"🔍 Event keys: {list(event.keys())}")
            
            for field in description_fields:
                field_value = event.get(field)
                if field_value:
                    description = field_value
                    print(f"✅ Found description in field '{field}': '{str(description)[:100]}...'")
                    break
                else:
                    print(f"❌ Field '{field}': {repr(field_value)}")
            
            if not description:
                print(f"⚠️ No description found for event '{event.get('title', event.get('name', 'Unknown'))}'")
                print(f"🔍 Full event data: {event}")
                description = 'No description available'
            
            # Process description through safe_excel_value
            safe_description = self.safe_excel_value(description)
            print(f"📝 Final description for Excel: '{safe_description}'")
            
            ws.cell(row=i, column=11, value=safe_description)  # Description
            ws.cell(row=i, column=12, value=self.safe_excel_value(event.get('notes', '')))  # Notes
            
            # Alternate row colors
            if i % 2 == 0:
                for col in range(1, 13):
                    cell = ws.cell(row=i, column=col)
                    if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == 'FFFFFFFF':
                        cell.fill = PatternFill(
                            start_color=self.colors['light'], 
                            end_color=self.colors['light'], 
                            fill_type='solid'
                        )
            
            # Color code paid status
            paid_cell = ws.cell(row=i, column=10)
            if paid_status == 'Paid':
                paid_cell.fill = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')  # Light green
                paid_cell.font = Font(color='FF006100')  # Dark green
            elif paid_status == 'Pending':
                paid_cell.fill = PatternFill(start_color='FFFFEB9C', end_color='FFFFEB9C', fill_type='solid')  # Light yellow
                paid_cell.font = Font(color='FF9C5700')  # Dark orange
            elif paid_status == 'Unpaid':
                paid_cell.fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')  # Light red
                paid_cell.font = Font(color='FF9C0006')  # Dark red
        
        # Set column widths - wider for better readability
        column_widths = [5, 12, 10, 10, 30, 25, 35, 15, 12, 12, 50, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Set row heights for better readability
        for row in range(2, len(trip_events) + 2):
            ws.row_dimensions[row].height = 30

    def format_cost(self, cost_info):
        """Format cost information for display"""
        if not cost_info or cost_info in ['N/A', 'NaN', None]:
            return 'Free'
        
        if isinstance(cost_info, (int, float)):
            if cost_info == 0:
                return 'Free'
            return f'${cost_info:.2f}'
        
        if isinstance(cost_info, str):
            if cost_info.lower() in ['free', 'n/a', 'nan', '0', '']:
                return 'Free'
            # Try to extract number from string
            import re
            numbers = re.findall(r'\d+\.?\d*', cost_info)
            if numbers:
                try:
                    amount = float(numbers[0])
                    return f'${amount:.2f}' if amount > 0 else 'Free'
                except ValueError:
                    pass
            return cost_info
        
        return str(cost_info)

    def get_paid_status(self, event):
        """Determine payment status of an event"""
        # Check various possible fields for payment status
        paid_fields = ['paid', 'isPaid', 'paymentStatus', 'status']
        
        for field in paid_fields:
            if field in event:
                value = event[field]
                if isinstance(value, bool):
                    return 'Paid' if value else 'Unpaid'
                elif isinstance(value, str):
                    value_lower = value.lower()
                    if value_lower in ['paid', 'completed', 'confirmed']:
                        return 'Paid'
                    elif value_lower in ['pending', 'processing']:
                        return 'Pending'
                    elif value_lower in ['unpaid', 'not paid', 'cancelled']:
                        return 'Unpaid'
        
        # Check if cost is 0 or free
        cost = event.get('cost', event.get('estimatedCost'))
        if cost in [0, '0', 'free', 'Free', 'N/A']:
            return 'Free'
        
        # Default status
        return 'TBD'

    def create_summary_sheet(self, wb, calendar_data, trip_data):
        """Create trip summary and notes sheet with much wider layout"""
        ws = wb.create_sheet("📝 Trip Summary & Notes")
        
        # Header - much wider
        ws.merge_cells('A1:J1')
        ws['A1'].value = f"📝 Trip Summary & Notes - {trip_data.get('name', 'Holiday Moo Trip')}"
        ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        # Trip description section - much wider
        ws.cell(row=3, column=1, value="📋 Trip Description:").font = Font(bold=True, size=14)
        ws.merge_cells('A4:J10')
        desc_cell = ws['A4']
        description = trip_data.get('description') or trip_data.get('notes') or 'No description provided. Add your trip details, objectives, and special notes here.'
        desc_cell.value = self.safe_excel_value(description)
        desc_cell.alignment = Alignment(vertical='top', wrap_text=True)
        desc_cell.font = Font(size=11)
        
        # Travel checklist section
        ws.cell(row=12, column=1, value="✅ Travel Checklist:").font = Font(bold=True, size=14)
        ws.merge_cells('A13:E25')
        checklist_cell = ws['A13']
        checklist_text = """□ Passport & Travel Documents
□ Flight/Transportation Tickets
□ Hotel Reservations
□ Travel Insurance
□ Currency Exchange
□ Phone & Chargers
□ Camera & Memory Cards
□ Medications
□ Weather-appropriate Clothing
□ Emergency Contact Information
□ Copies of Important Documents
□ Travel Adapters"""
        checklist_cell.value = checklist_text
        checklist_cell.alignment = Alignment(vertical='top', wrap_text=True)
        checklist_cell.font = Font(size=10)
        
        # Important information section
        ws.cell(row=12, column=6, value="📞 Important Information:").font = Font(bold=True, size=14)
        ws.merge_cells('F13:J25')
        info_cell = ws['F13']
        info_text = """Emergency Contacts:
• Local Emergency: [Add local emergency number]
• Embassy/Consulate: [Add contact info]
• Travel Insurance: [Add policy number]
• Bank/Credit Card: [Add contact numbers]

Important Addresses:
• Accommodation: [Add hotel/accommodation address]
• Airport: [Add airport information]
• Meeting Points: [Add important locations]

Local Information:
• Currency: [Add local currency info]
• Tipping: [Add tipping customs]
• Language: [Add key phrases]
• Time Zone: [Add time difference]"""
        info_cell.value = info_text
        info_cell.alignment = Alignment(vertical='top', wrap_text=True)
        info_cell.font = Font(size=10)
        
        # Notes section - full width
        ws.cell(row=27, column=1, value="📝 Additional Notes & Reminders:").font = Font(bold=True, size=14)
        ws.merge_cells('A28:J40')
        notes_cell = ws['A28']
        notes_text = """Add your personal travel notes here:

• Special dietary requirements or preferences
• Shopping lists and souvenirs to buy
• Photo spots and must-see locations
• Local customs and etiquette to remember
• Weather considerations and backup plans
• Transportation tips and local navigation
• Restaurant recommendations and food to try
• Cultural events or festivals during your visit
• Language tips and useful phrases
• Safety considerations and precautions"""
        notes_cell.value = notes_text
        notes_cell.alignment = Alignment(vertical='top', wrap_text=True)
        notes_cell.font = Font(size=10)
        
        # Generated info - full width
        ws.merge_cells('A42:J42')
        generated_cell = ws['A42']
        generated_cell.value = f"Generated by Holiday Moo 🏖️ on {datetime.now().strftime('%B %d, %Y at %H:%M')} | Visit us at holidaymoo.com"
        generated_cell.font = Font(italic=True, size=9, color='FF666666')
        generated_cell.alignment = Alignment(horizontal='center')
        
        # Set much wider column widths
        column_widths = [15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Set row heights for better spacing
        for row in [4, 13, 28]:
            ws.row_dimensions[row].height = 200

    def parse_datetime(self, date_string):
        """Parse various datetime formats as timezone-naive (local time)"""
        if not date_string:
            return datetime.now()
            
        # Remove timezone info completely - treat all times as local
        clean_date = str(date_string)
        
        # Remove common timezone indicators
        clean_date = clean_date.replace('Z', '')  # UTC indicator
        clean_date = clean_date.replace('+00:00', '')  # UTC offset
        clean_date = clean_date.replace('T00:00:00.000', '')  # Midnight with milliseconds
        
        # Remove milliseconds if present
        if '.' in clean_date:
            clean_date = clean_date.split('.')[0]
        
        # Try different formats
        formats = [
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d',
            '%Y/%m/%d %H:%M:%S',
            '%Y/%m/%d',
            '%m/%d/%Y %H:%M:%S',
            '%m/%d/%Y'
        ]
        
        for fmt in formats:
            try:
                parsed_dt = datetime.strptime(clean_date, fmt)
                print(f"🕐 Parsed '{date_string}' -> '{clean_date}' as {parsed_dt} (format: {fmt})")
                return parsed_dt
            except ValueError:
                continue
        
        # If all formats fail, try parsing just the date part
        try:
            date_part = clean_date[:10] if len(clean_date) >= 10 else clean_date
            parsed_dt = datetime.strptime(date_part, '%Y-%m-%d')
            print(f"🕐 Parsed date part '{date_string}' -> '{date_part}' as {parsed_dt}")
            return parsed_dt
        except ValueError:
            # Last resort - return current date
            print(f"⚠️ Could not parse datetime '{date_string}', using current time")
            return datetime.now()

    def safe_excel_value(self, value):
        """Convert any value to Excel-safe format"""
        if value is None:
            return ""
        elif isinstance(value, (str, int, float, bool)):
            return value
        elif isinstance(value, dict):
            # Convert dict to readable string
            if 'name' in value:
                # Location object with name
                result = value['name']
                if 'address' in value:
                    result += f" ({value['address']})"
                return result
            else:
                # Generic dict - convert to key: value pairs
                return ", ".join([f"{k}: {v}" for k, v in value.items() if isinstance(v, (str, int, float, bool))])
        elif isinstance(value, list):
            # Convert list to comma-separated string
            return ", ".join([str(item) for item in value if isinstance(item, (str, int, float, bool))])
        else:
            # Convert anything else to string
            return str(value)

    def get_trip_events(self, events, start_date, end_date):
        """Filter events for the trip date range"""
        trip_events = []
        for event in events:
            try:
                event_date = self.extract_date_only(event.get('startTime', '2025-01-01'))
                if start_date.date() <= event_date <= end_date.date():
                    trip_events.append(event)
                    print(f"✅ Event '{event.get('title', 'Unknown')}' on {event_date} included in trip")
                else:
                    print(f"❌ Event '{event.get('title', 'Unknown')}' on {event_date} outside trip range {start_date.date()} - {end_date.date()}")
            except Exception as e:
                print(f"Warning: Could not parse event date {event.get('startTime', 'unknown')}: {e}")
                continue
        return trip_events

    def generate_excel(self, calendar_data, trip_data):
        """Main method to generate Excel file"""
        wb = self.create_workbook(calendar_data, trip_data)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Encode to base64
        excel_b64 = base64.b64encode(excel_buffer.getvalue()).decode('utf-8')
        
        # Generate filename
        trip_name = trip_data['name'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"HolidayMoo_{trip_name}_{timestamp}.xlsx"
        
        return {
            'success': True,
            'filename': filename,
            'data': excel_b64,
            'size': len(excel_buffer.getvalue())
        }

# Flask routes
@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy', 'service': 'Holiday Moo Local Export'})

@app.route('/export-trip', methods=['POST'])
def export_trip():
    try:
        data = request.get_json()
        
        if not data or 'calendarData' not in data or 'tripData' not in data:
            return jsonify({'success': False, 'error': 'Missing required data'}), 400
        
        calendar_data = data['calendarData']
        trip_data = data['tripData']
        
        # Debug logging
        print(f"📊 Processing export for trip: {trip_data.get('name', 'Unknown')}")
        print(f"📅 Trip dates: {trip_data.get('startDate')} to {trip_data.get('endDate')}")
        print(f"📋 Total events: {len(calendar_data.get('events', []))}")
        print(f"📄 Calendar data keys: {list(calendar_data.keys())}")
        print(f"🎯 Trip data keys: {list(trip_data.keys())}")
        
        # Sample event for debugging
        if calendar_data.get('events'):
            sample_event = calendar_data['events'][0]
            print(f"🔍 Sample event keys: {list(sample_event.keys())}")
            print(f"🔍 Sample event startTime RAW: {repr(sample_event.get('startTime', 'No startTime'))}")
            print(f"🔍 Sample event endTime RAW: {repr(sample_event.get('endTime', 'No endTime'))}")
            
            # Test time extraction on sample event
            generator = HolidayMooExcelGenerator()
            if sample_event.get('startTime'):
                hour, minute = generator.extract_local_time(sample_event['startTime'])
                print(f"🕐 Extracted start time: {hour:02d}:{minute:02d}")
            
            print(f"🔍 Sample event title: {sample_event.get('title', 'No title')}")
            print(f"🔍 Sample event name: {sample_event.get('name', 'No name')}")
            print(f"🔍 Sample event description: {sample_event.get('description', 'No description')}")
            print(f"🔍 Sample event location type: {type(sample_event.get('location'))}")
            print(f"🔍 Sample event location: {sample_event.get('location', 'No location')}")
            print(f"🔍 Sample event cost: {sample_event.get('cost', sample_event.get('estimatedCost', 'No cost'))}")
            print(f"🔍 Full sample event: {sample_event}")
        
        # Generate Excel
        generator = HolidayMooExcelGenerator()
        result = generator.generate_excel(calendar_data, trip_data)
        
        print(f"✅ Excel generated successfully: {result['filename']}")
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    print("🏖️ Holiday Moo Local Export Service Starting...")
    print("📊 Beautiful Excel Dashboard Generator Ready!")
    print("🌐 Server running on http://localhost:5001")
    print("📅 Calendar-focused export with professional formatting")
    app.run(host='0.0.0.0', port=5001, debug=True)